from operator import eq, ne
from pprint import pprint
import xlrd
import json, yaml, re
from collections import OrderedDict
from generators.BlankNone import *
from openpyxl import load_workbook


def convertToBoolIfNeeded(variable):
	if type(variable) == str and re.match(r'(?i)(True|False)', variable.strip()):
		variable = True if re.match(r'(?i)true', variable.strip()) else False
	return variable

def convertToList(key, value, types):	
	keys_that_are_lists = ["mlagInterfaces", "uplinkSwitches", "uplinkInterfaces", "uplinkSwitchInterfaces"]
	
	if key in keys_that_are_lists:
		value = [v.strip() for v in value.split(",") if v != ""]
	return value

def consolidateNodeGroups(node_groups):
	# node_group_level_vars = ["bgp_as", "platform", "filter", "parent_l3leafs", "uplink_interfaces"]
	node_group_level_vars = ["filter", "uplink_switches"]
	groups_names = node_groups.keys()
	new_group_vars = {}
	for group, nodes in node_groups.items():
		new_group_vars[group] = {}
		for node, details in nodes.items():               
			if len(details.keys()) > 1:
				host1_vars = details[list(details.keys())[0]]
				host2_vars = details[list(details.keys())[1]]
				for variable, value in host1_vars.items():
					if variable in host2_vars.keys() and host1_vars[variable] == host2_vars[variable]:                        
							new_group_vars[group][variable] = value
					elif variable in node_group_level_vars:
						new_group_vars[group][variable] = value 
			else:
					for variable, value in details[list(details.keys())[0]].items():
						if variable in node_group_level_vars:
							new_group_vars[group][variable] = value 


	# print(json.dumps(new_group_vars, indent=2))
	for group, variables in new_group_vars.items():
		for variable, value in variables.items():
			node_groups[group][variable] = value
			for variable_dict in node_groups[group]["nodes"].values():
				del(variable_dict[variable])
	return node_groups 

def parseL2LeafInfo(inventory_file):
	l2leaf = L2Leaf()
	l2leafDetail = L2LeafDetail()
	l2_yaml = {}
	configuration_variable_mappers = {
		l2leafDetail.platform: "platform", 
		l2leafDetail.uplinkSwitches:"uplink_switches", 
		l2leafDetail.uplinkInterfaces: "uplink_interfaces",
		l2leafDetail.mlagInterfaces:"mlag_interfaces", 
		l2leafDetail.mlag: "mlag", 
		l2leafDetail.mlagPeerIpv4Pool: "mlag_peer_ipv4_pool", 
		l2leafDetail.mlagPeerL3Ipv4Pool: "mlag_peer_l3_ipv4_pool", 
		l2leafDetail.virtualRouterMacAddress: "virtual_router_mac_address", 
		l2leafDetail.spanningTreeMode:"spanning_tree_mode", 
		l2leafDetail.spanningTreePriority:"spanning_tree_priority"
	}
	l2_leaf_info = {}
	workbook = xlrd.open_workbook(inventory_file)
	inventory_worksheet = workbook.sheet_by_name("L2 Leaf Info")
	node_groups = {}
	first_row = [] # The row where we stock the name of the column
	for col in range(inventory_worksheet.ncols):
		first_row.append( inventory_worksheet.cell_value(0,col) )
	# transform the workbook to a list of dictionaries
	for row in range(1, inventory_worksheet.nrows):
		l2_leaf_info = {}
		for col in range(inventory_worksheet.ncols):
			l2_leaf_info[first_row[col]]=inventory_worksheet.cell_value(row,col)
		hostname = l2_leaf_info[l2leaf.hostname]
		node_details = {}
		node_details["id"] = int(l2_leaf_info[l2leaf.id])
		node_details["mgmt_ip"] = l2_leaf_info[l2leaf.managementIp]

		if ne(l2_leaf_info[l2leaf.tenats], "") or ne(l2_leaf_info["Tags"], "") :
			node_details["filter"] = {}
			if ne(l2_leaf_info[l2leaf.tenats], ""):
				node_details["filter"]["tenants"] = [tenant.strip() for tenant in l2_leaf_info[l2leaf.tenats].split(",") if tenant != ""]
			if ne(l2_leaf_info[l2leaf.tags], ""):
				node_details["filter"]["tags"] = [tag.strip() for tag in l2_leaf_info[l2leaf.tags].split(",") if tag != ""]
		
		optional_params = {}
		# optional_params["platform"] = str(l2_leaf_info["Platform"]) if l2_leaf_info["Platform"] != "" else None
		optional_params["uplink_switches"] = [spine.strip() for spine in l2_leaf_info[l2leaf.uplinkSwitches].split(",") if spine] if l2_leaf_info[l2leaf.uplinkSwitches] != "" else None
		# optional_params["uplink_interfaces"] = [uplink_iface.strip() for uplink_iface in l2_leaf_info["Uplink Interfaces"].split(",") if uplink_iface] if l2_leaf_info["Uplink Interfaces"] != "" else None
		optional_params["uplink_switch_interfaces"] = [uplink_iface.strip() for uplink_iface in l2_leaf_info[l2leaf.uplinkSwitchInterfaces].split(",") if uplink_iface] if l2_leaf_info[l2leaf.uplinkSwitchInterfaces] != "" else None
		# optional_params["mlag_interfaces"] = [iface.strip() for iface in l2_leaf_info["MLAG Interfaces"].split(",") if iface] if l2_leaf_info["MLAG Interfaces"] != "" else None
		for k, v in optional_params.items():
			if v is not None:
				v = int(v) if type(v) == float else v
				node_details[k] = v

		if l2_leaf_info[l2leaf.containerName] not in node_groups.keys():
			node_groups[l2_leaf_info[l2leaf.containerName]] = {"nodes": {hostname: node_details}}
		else:
			node_groups[l2_leaf_info[l2leaf.containerName]]["nodes"][hostname] = node_details

	# print(yaml.dump(node_groups))
	#parse default values
	l2_defaults_worksheet = workbook.sheet_by_name(l2leafDetail.sheetName)
	defaults = {}
	# transform the workbook to a list of dictionaries
	for row in range(1, l2_defaults_worksheet.nrows):
		k, v = l2_defaults_worksheet.cell_value(row,0), l2_defaults_worksheet.cell_value(row,1)
		if k in configuration_variable_mappers.keys() and v is not None and v != "":
			v = convertToList(k, v, "L2")
			v = convertToBoolIfNeeded(v)
			v = int(v) if type(v) == float else v
			defaults[configuration_variable_mappers[k]] = v

	# print(json.dumps(defaults, indent=2))
	l2_yaml["defaults"] = defaults
	l2_yaml["node_groups"] = consolidateNodeGroups(node_groups)

	return l2_yaml

def parseL3LeafInfo(inventory_file, excelVar):
	
	l3_yaml = {}
	configuration_variable_mappers = {
		"platform", "loopback_ipv4_pool", "loopback_ipv4_offset", "vtep_loopback_ipv4_pool", "uplink_interfaces", "uplink_switches", "uplink_ipv4_pool", "mlag_interfaces", "mlag_port_channel_id", "mlag_peer_ipv4_pool", "mlag_peer_l3_ipv4_pool", "virtual_router_mac_address", "spanning_tree_mode", "spanning_tree_priority", "prefix_name", "prefix_sequence_number", "prefix_action", "route_map_name", "route_map_type", "route_map_sequence", "route_map_match"
	}

	workbook = load_workbook(filename=inventory_file, read_only=True, data_only=True)
	inventory_worksheet = workbook[excelVar["leaf"]["sheet"]]
	spinePrefix = excelVar["spine"]["prefix"]
	leafPrefix = excelVar["leaf"]["prefix"]
	leafBgpAsCol = excelVar["leaf"]["props"]["bgp"]["col"]
	leafHostnameCol = excelVar["leaf"]["props"]["hostname"]["col"]
	portmapSheet = excelVar["leafUplink"]["sheet"] 
	startSwitch = excelVar["leafUplink"]["start"]["switch"]
	startPort =  excelVar["leafUplink"]["start"]["port"]
	endSwitch =  excelVar["leafUplink"]["end"]["switch"]
	endPort =  excelVar["leafUplink"]["end"]["port"]
 
	portMap = {}
	mlagMap = {}
	node_groups = {}
	inventory_worksheet = workbook[portmapSheet]

	## spine, leaf uplink 체크
	for row in inventory_worksheet.iter_rows():
		for cell in row:
  		# print(cell)
			if cell.value:
				if eq(cell.coordinate, startSwitch + str(cell.row)):
					p = re.compile(spinePrefix)
					if (p.match(str(cell.value))):
						leaf = inventory_worksheet[endSwitch + str(cell.row)].value
						if not leaf in portMap:
							portMap.setdefault(
								leaf,  {
									"uplinkInterfaces": [inventory_worksheet[endPort + str(cell.row)].value],
									"remoteInterfaces": [inventory_worksheet[startPort + str(cell.row)].value]
								}
							)
					
						else:
							portMap[leaf]["uplinkInterfaces"].append(inventory_worksheet[endPort + str(cell.row)].value)
							portMap[leaf]["remoteInterfaces"].append(inventory_worksheet[startPort + str(cell.row)].value)



	## mlag 체크
	for row in inventory_worksheet.iter_rows():
		for cell in row:
  		# print(cell)
			if cell.value:
				if eq(cell.coordinate, startSwitch + str(cell.row)):
					p = re.compile(leafPrefix)
					if (p.match(str(cell.value))):
						mlag1 = inventory_worksheet[startSwitch + str(cell.row)].value
						mlag2 = inventory_worksheet[endSwitch + str(cell.row)].value
						mlag = mlag1 + "|" + mlag2
						if not mlag in mlagMap:
							mlagMap.setdefault(
								mlag,  {
									"switches": [mlag1, mlag2],
									"mlag_interfaces": [inventory_worksheet[endPort + str(cell.row)].value]
								}
							)
					
						else:
							mlagMap[mlag]["mlag_interfaces"].append(inventory_worksheet[endPort + str(cell.row)].value)

	inventory_worksheet = workbook[excelVar["leaf"]["sheet"]]
 
	for row in inventory_worksheet.iter_rows():
		for cell in row:
			# print(cell)
			if cell.value:
				if eq(cell.coordinate, leafHostnameCol + str(cell.row)):
					## leaf
					p = re.compile(leafPrefix)
					# print(cell.value)
					if (p.match(str(cell.value))):
						codi = excelVar["leaf"]["props"]["id"]["col"] + str(cell.row)
						id = inventory_worksheet[codi].value
						codi = excelVar["leaf"]["props"]["hostname"]["col"] + str(cell.row)
						hostname = inventory_worksheet[codi].value
						codi = excelVar["leaf"]["props"]["mgmt"]["col"] + str(cell.row)
						mgmtIp = inventory_worksheet[codi].value
						node_details = {}
						node_details["id"] = int(id)
						node_details["mgmt_ip"] = mgmtIp

						if ne(portMap[hostname]["remoteInterfaces"], ""):
							node_details["uplink_switch_interfaces"] = portMap[hostname]["remoteInterfaces"]
					
						optional_params = {}
			
						bgpAs = inventory_worksheet[leafBgpAsCol + str(cell.row)].value
						optional_params["bgp_as"] = int(bgpAs) if bgpAs != "" else None

						## mlag						
						mlagExsits = False
						for item in mlagMap:
							if hostname in mlagMap[item]["switches"]:
								optional_params["mlag_interfaces"] = mlagMap[item]["mlag_interfaces"]
								mlagExsits = True
        
						if not mlagExsits:
							optional_params["mlag"] = False
        
						for k, v in optional_params.items():
							if v is not None:
								v = int(v) if type(v) == float else v
								node_details[k] = v
        
						containerName =  inventory_worksheet[leafContainerCol + str(cell.row)].value

						if containerName not in node_groups.keys():
							node_groups[containerName] = {"nodes": {hostname: node_details}}
						else:
							node_groups[containerName]["nodes"][hostname] = node_details

	# pprint(node_groups)
	# exit()
	# print(yaml.dump(node_groups))
	#parse default values
	defaults = {}
	bgp_defaults = {}
	# transform the workbook to a list of dictionaries
	for item in excelVar["leafInfo"]:
		v = getExcelSheetValue(workbook, excelVar["leafInfo"][item])
		if excelVar["leafInfo"][item]["mapping"] in configuration_variable_mappers and v is not None and v != "":
			v = convertToList(item, v, "L3")
			v = convertToBoolIfNeeded(v)
			v = int(v) if type(v) == float else v
			defaults[excelVar["leafInfo"][item]["mapping"]] = v


	l3_yaml["defaults"] = defaults
	l3_yaml["node_groups"] = consolidateNodeGroups(node_groups)
	# pprint(defaults)
	##### Prefix 처리 S #####
	if ne("", l3_yaml["defaults"]["prefix_name"]) and ne("", l3_yaml["defaults"]["prefix_sequence_number"]) and ne("", l3_yaml["defaults"]["prefix_action"]):
		
		prefix_name = [iface.strip() for iface in l3_yaml["defaults"]["prefix_name"].split(",") if iface] if l3_yaml["defaults"]["prefix_name"] != "" else None
		prefix_sequence_number = [iface.strip() for iface in str(l3_yaml["defaults"]["prefix_sequence_number"]).split(",") if iface] if l3_yaml["defaults"]["prefix_sequence_number"] != "" else None
		prefix_action = [iface.strip() for iface in l3_yaml["defaults"]["prefix_action"].split(",") if iface] if l3_yaml["defaults"]["prefix_action"] != "" else None

		item = []
		for i in range(0, len(prefix_name)):
			item.append({
				"name" : prefix_name[i],
				"sequence_numbers" : [{
					"sequence": prefix_sequence_number[i],
					"action": prefix_action[i]
				}]
			})

		l3_yaml["defaults"]["prefix_lists"] = item


	del(l3_yaml["defaults"]["prefix_name"])
	del(l3_yaml["defaults"]["prefix_sequence_number"])
	del(l3_yaml["defaults"]["prefix_action"])
	##### Prefix 처리 E #####

	##### Route Map 처리 S #####
	if ne("", l3_yaml["defaults"]["route_map_name"]) and ne("", l3_yaml["defaults"]["route_map_type"]) and ne("", l3_yaml["defaults"]["route_map_sequence"]) and ne("", l3_yaml["defaults"]["route_map_match"]):
		l3_yaml["defaults"]["route_maps"] = [{
				"name" : l3_yaml["defaults"]["route_map_name"],
				"sequence_numbers" : [{
					"sequence": l3_yaml["defaults"]["route_map_sequence"],
					"type": l3_yaml["defaults"]["route_map_type"],
					"match": [l3_yaml["defaults"]["route_map_match"]]
				}]
			}]
	del(l3_yaml["defaults"]["route_map_name"])
	del(l3_yaml["defaults"]["route_map_type"])
	del(l3_yaml["defaults"]["route_map_sequence"])
	del(l3_yaml["defaults"]["route_map_match"])
	##### Route Map E #####
	
	# BGP default 세팅
	l3_yaml["defaults"]["bgp_defaults"] = parseLeafBGPDefaults(inventory_file, excelVar)

	return l3_yaml

def parseL3LeafBGPDefaults(inventory_file):
	#parse default values
	l3leafDetail = L3LeafDetail()
	configuration_variable_mappers = {"BGP wait-install": "wait_install", "BGP distance setting":"distance_setting", "BGP default ipv4-unicast": "ipv4_unicast"}
	l3_leaf_info = {}
	workbook = xlrd.open_workbook(inventory_file)
	l3_defaults_worksheet = workbook.sheet_by_name(l3leafDetail.sheetName)
	bgp_defaults = {}
	# transform the workbook to a list of dictionaries
	for row in range(1, l3_defaults_worksheet.nrows):
		k, v = l3_defaults_worksheet.cell_value(row,0), l3_defaults_worksheet.cell_value(row,1)
		if k in configuration_variable_mappers.keys() and v is not None and v != "":
			v = convertToBoolIfNeeded(v)
			bgp_defaults[configuration_variable_mappers[k]] = v
	bgp_defaults_list = []
	config_values = {
		"wait_install":
			{
				True: "update wait-install",
				False: None
			},
		"ipv4_unicast":
		{
			True: "bgp default ipv4-unicast",
			False: "no bgp default ipv4-unicast"
		}
	}
	for k, v in bgp_defaults.items():
		if k in config_values.keys():
			v = config_values[k][bool(v)]
		if v is not None:
			bgp_defaults_list.append(v)
	return bgp_defaults_list

# Spine Switches
def parseSpineInfo(inventory_file, excelVar):
	spine_yaml = {"defaults": {}, "nodes": {}}

	configuration_variable_mappers = {
		"platform", "leaf_as_range", "bgp_as", "loopback_ipv4_pool", "mlag_peer_ipv4_pool", "mlag_peer_l3_ipv4_pool", "super_spine", "peer_filter_name", "peer_filter_sequence_number", "peer_filter_match", "prefix_name", "prefix_sequence_number", "prefix_action", "route_map_name", "route_map_type", "route_map_sequence", "route_map_match"
	}
	
	workbook = load_workbook(filename=inventory_file, read_only=True, data_only=True)
	inventory_worksheet = workbook[excelVar["spine"]["sheet"]]

	spinePrefix = excelVar["spine"]["prefix"]
	spineHostnameCol = excelVar["spine"]["props"]["hostname"]["col"]

	for row in inventory_worksheet.iter_rows():
		for cell in row:
			# print(cell)
			if cell.value:
				if eq(cell.coordinate, spineHostnameCol + str(cell.row)):
					p = re.compile(spinePrefix)
					if (p.match(str(cell.value))):
						codi = excelVar["spine"]["props"]["id"]["col"] + str(cell.row)
						id = inventory_worksheet[codi].value
						codi = excelVar["spine"]["props"]["hostname"]["col"] + str(cell.row)
						hostname = inventory_worksheet[codi].value
						codi = excelVar["spine"]["props"]["mgmt"]["col"] + str(cell.row)
						mgmtIp = inventory_worksheet[codi].value
						node_details = {}
						node_details["id"] = int(id)
						node_details["mgmt_ip"] = mgmtIp

						spine_yaml["nodes"][hostname] = node_details

	#parse default values
	for item in excelVar["spineInfo"]:
		v = getExcelSheetValue(workbook, excelVar["spineInfo"][item])
		if excelVar["spineInfo"][item]["mapping"] in configuration_variable_mappers and v is not None and v != "":
			v = convertToBoolIfNeeded(v)
			v = int(v) if type(v) == float else v
			spine_yaml["defaults"][excelVar["spineInfo"][item]["mapping"]] = v

	##### Peer Filter 처리 S #####
	if ne("", spine_yaml["defaults"]["peer_filter_name"]) and ne("", spine_yaml["defaults"]["peer_filter_sequence_number"]) and ne("", spine_yaml["defaults"]["peer_filter_match"]):
		spine_yaml["defaults"]["peer_filters"] = [{
				"name" : spine_yaml["defaults"]["peer_filter_name"],
				"sequence_numbers" : [{
					"sequence": spine_yaml["defaults"]["peer_filter_sequence_number"],
					"match": spine_yaml["defaults"]["peer_filter_match"]
				}]
			}]
		
	del(spine_yaml["defaults"]["peer_filter_name"])
	del(spine_yaml["defaults"]["peer_filter_sequence_number"])
	del(spine_yaml["defaults"]["peer_filter_match"])
	##### Peer Filter 처리 E #####

	##### Prefix 처리 S #####
	if ne("", spine_yaml["defaults"]["prefix_name"]) and ne("", spine_yaml["defaults"]["prefix_sequence_number"]) and ne("", spine_yaml["defaults"]["prefix_action"]):
		spine_yaml["defaults"]["prefix_lists"] = [{
				"name" : spine_yaml["defaults"]["prefix_name"],
				"sequence_numbers" : [{
					"sequence": spine_yaml["defaults"]["prefix_sequence_number"],
					"action": spine_yaml["defaults"]["prefix_action"]
				}]
			}]
	del(spine_yaml["defaults"]["prefix_name"])
	del(spine_yaml["defaults"]["prefix_sequence_number"])
	del(spine_yaml["defaults"]["prefix_action"])
	##### Prefix 처리 E #####

	##### Route Map 처리 S #####
	if ne("", spine_yaml["defaults"]["route_map_name"]) and ne("", spine_yaml["defaults"]["route_map_type"]) and ne("", spine_yaml["defaults"]["route_map_sequence"]) and ne("", spine_yaml["defaults"]["route_map_match"]):
		spine_yaml["defaults"]["route_maps"] = [{
				"name" : spine_yaml["defaults"]["route_map_name"],
				"sequence_numbers" : [{
					"sequence": spine_yaml["defaults"]["route_map_sequence"],
					"type": spine_yaml["defaults"]["route_map_type"],
					"match": [spine_yaml["defaults"]["route_map_match"]]
				}]
			}]
	del(spine_yaml["defaults"]["route_map_name"])
	del(spine_yaml["defaults"]["route_map_type"])
	del(spine_yaml["defaults"]["route_map_sequence"])
	del(spine_yaml["defaults"]["route_map_match"])
	##### Route Map E #####

	spine_yaml["defaults"]["bgp_defaults"] = parseSpineBGPDefaults(inventory_file, excelVar)

	return spine_yaml

def parseLeafBGPDefaults(inventory_file, excelVar):
  #parse default values
	# BGP default 세팅
	configuration_variable_mappers = {
		"bgp_distance_setting",
		"bgp_default_ipv4_unicast",
		"bgp_graceful_restart_time",
		"bgp_graceful_restart"
	}
	bgp_defaults = {}
	# transform the workbook to a list of dictionaries
	workbook = load_workbook(filename=inventory_file, read_only=True, data_only=True)

	bgp_defaults = {}
	# transform the workbook to a list of dictionaries
	for item in excelVar["leafInfo"]:
		v = getExcelSheetValue(workbook, excelVar["leafInfo"][item])
		if excelVar["leafInfo"][item]["mapping"] in configuration_variable_mappers and v is not None and v != "":
			v = convertToBoolIfNeeded(v)
			bgp_defaults[excelVar["leafInfo"][item]["mapping"]] = v
   
		k, v = item, v
		if k in configuration_variable_mappers and v is not None and v != "":
			v = convertToBoolIfNeeded(v)
			bgp_defaults[k] = v

			if eq("bgp_graceful_restart_time", excelVar["leafInfo"][k]["mapping"]):
				# print("graceful-restart restart-time !!")
				v = bgp_defaults[k]
				v = int(v) if type(v) == float else v
				v = str(v)
				bgp_defaults[k] = "graceful-restart restart-time " + v

	# pprint(bgp_defaults)
	
	bgp_defaults_list = []
	config_values = {
		"bgp_default_ipv4_unicast":
			{
				True: "bgp default ipv4-unicast",
				False: "no bgp default ipv4-unicast"
			},
		"bgp_graceful_restart":
			{
				True: "graceful-restart",
				False: None
			}
	}

	for k, v in bgp_defaults.items():
		if k in config_values.keys():
			v = config_values[k][bool(v)]
		if v is not None:
			bgp_defaults_list.append(v)
			
  
	return bgp_defaults_list

def parseSpineBGPDefaults(inventory_file, excelVar):
	#parse default values
	
	configuration_variable_mappers = {
		"update_wait_for_convergence", "wait_install", "distance_setting","ipv4_unicast"
	}
	workbook = load_workbook(filename=inventory_file, read_only=True, data_only=True)

	bgp_defaults = {}
	# transform the workbook to a list of dictionaries
	for item in excelVar["spineInfo"]:
		v = getExcelSheetValue(workbook, excelVar["spineInfo"][item])
		if excelVar["spineInfo"][item]["mapping"] in configuration_variable_mappers and v is not None and v != "":
			v = convertToBoolIfNeeded(v)
			bgp_defaults[excelVar["spineInfo"][item]["mapping"]] = v

			# if eq(item, "bgp_graceful_restart_time"):
			# 	# print("graceful-restart restart-time !!")
			# 	v = bgp_defaults[configuration_variable_mappers[k]]
			# 	v = int(v) if type(v) == float else v
			# 	v = str(v)
			# 	bgp_defaults[configuration_variable_mappers[k]] = "graceful-restart restart-time " + v
	
	bgp_defaults_list = []
	config_values = {
		"update_wait_for_convergence":
			{
				True: "update wait-for-convergence",
				False: None
			},
		"wait_install":
			{
				True: "update wait-install",
				False: None
			},
		"ipv4_unicast":
		{
			True: "bgp default ipv4-unicast",
			False: "no bgp default ipv4-unicast"
		}
	}
	for k, v in bgp_defaults.items():
		if k in config_values.keys():
			v = config_values[k][bool(v)]
		if v is not None:
			bgp_defaults_list.append(v)
   
	return bgp_defaults_list

def parseGeneralVariables(inventory_file, excelVar):
	general_yaml = {}
	
	#parse default values
	workbook = load_workbook(filename=inventory_file, read_only=False, data_only=True)
	general_yaml = {}

	for item in excelVar["general"]:
		v = getExcelSheetValue(workbook, excelVar["general"][item])
		v = convertToBoolIfNeeded(v)
		v = v if v != "" else None
		v = int(v) if type(v) == float else v

		if (eq(item, "fabricName")):
			v = v + "_FABRIC"

		general_yaml[excelVar["general"][item]["mapping"]] = v
	
	# general_yaml["bgp_peer_groups"] = {
	# 	"IPv4_UNDERLAY_PEERS":{"password": general_yaml["bgp_ipv4_password"]},
	# 	"EVPN_OVERLAY_PEERS":{"password": general_yaml["bgp_evpn_password"]},
	# 	"MLAG_IPv4_UNDERLAY_PEER":{"password": general_yaml["bgp_mlag_ipv4_password"]}
	# }
	
	general_yaml["bgp_peer_groups"] = {
		"IPv4_UNDERLAY_PEERS":{ "name": general_yaml["bgp_ipv4_name"], "bgp_listen_range_prefix": general_yaml["bgp_ipv4_prefix"], "peer_filter": general_yaml["bgp_ipv4_filter"], "password": general_yaml["bgp_ipv4_password"], "remote_as": general_yaml["bgp_ipv4_remoteas"], "maximum_routes": general_yaml["bgp_ipv4_maximum_routes"]},
		"EVPN_OVERLAY_PEERS":{ "name": general_yaml["bgp_evpn_name"], "bgp_listen_range_prefix": general_yaml["bgp_evpn_prefix"], "peer_filter": general_yaml["bgp_evpn_filter"], "remote_as": general_yaml["bgp_evpn_remoteas"], "password": general_yaml["bgp_evpn_password"]}
	}

	del(general_yaml["bgp_ipv4_name"])
	del(general_yaml["bgp_ipv4_prefix"])
	del(general_yaml["bgp_ipv4_filter"])
	del(general_yaml["bgp_ipv4_password"])
	del(general_yaml["bgp_ipv4_remoteas"])
	del(general_yaml["bgp_ipv4_maximum_routes"])
	del(general_yaml["bgp_evpn_name"])
	del(general_yaml["bgp_evpn_prefix"])
	del(general_yaml["bgp_evpn_filter"])
	del(general_yaml["bgp_evpn_password"])
	del(general_yaml["bgp_evpn_remoteas"])
 
	general_yaml["bfd_multihop"] = {
		"interval": general_yaml["bfd_interval"], 
		"min_rx": general_yaml["bfd_min_rx"], 
		"multiplier": general_yaml["bfd_multiplier"]
	}

	del(general_yaml["bfd_interval"])
	del(general_yaml["bfd_min_rx"])
	del(general_yaml["bfd_multiplier"])

	return general_yaml

def generateGroupVarsFabric(file_location, excelVar):

	fabric_name = parseGeneralVariables(file_location, excelVar)
	# fabric_name["super_spine"] = parseSuperSpineInfo(file_location)
	fabric_name["spine"] = parseSpineInfo(file_location, excelVar)
	fabric_name["l3leaf"] = parseL3LeafInfo(file_location, excelVar)
	# fabric_name["l2leaf"] = parseL2LeafInfo(file_location, excelVar)
	# fabric_name["leaf_bgp_defaults"] = parseL3LeafBGPDefaults(file_location, excelVar)

	return fabric_name
