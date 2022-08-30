from multiprocessing.sharedctypes import Value
from pickle import NONE
import xlrd, re, ipaddress
import yaml
from generators.BlankNone import *
from openpyxl import load_workbook
from operator import eq

def convertToBoolIfNeeded(variable):
	if type(variable) == str and re.match(r'(?i)(True|False)', variable.strip()):
		variable = True if re.match(r'(?i)true', variable.strip()) else False
	return variable

def getFabricName(inventory_file, excelVar):
  workbook = load_workbook(filename=inventory_file, read_only=False, data_only=True)
  return getExcelSheetValue(workbook, excelVar["all"]["fabricName"])

def getCVPAddresses(inventory_file):
	workbook = load_workbook(filename=inventory_file, read_only=False, data_only=True)
	info_worksheet = workbook.sheet_by_name("General Configuration Details")
	# transform the workbook to a list of dictionaries
	for row in range(1, info_worksheet.nrows):
		k, v = info_worksheet.cell_value(row,0), info_worksheet.cell_value(row,1)
		if k == "CVP IP Addresses":
			return [ip.strip() for ip in v.split(",") if ip != ""]
	return None

def getCVPInventory(inventory_file):
	cvp_addresses = getCVPAddresses(inventory_file)
	cvp_dict = {"hosts": {}}
	cvp_node_names = ["cvp_primary", "cvp_secondary", "cvp_tertiary"]
	for i, address in enumerate(cvp_addresses):
		cvp_dict["hosts"][cvp_node_names[i]] = {
			"ansible_httpapi_host": address,
			"ansible_host": address
		}
		break
	return cvp_dict

def parseSpineInfo(inventory_file, excelVar):
	'''
	'''
	spines_info = {"vars": {"type": "spine"}, "hosts": {}}
	workbook = load_workbook(filename=inventory_file, read_only=True, data_only=True)
	inventory_worksheet = workbook[excelVar["spine"]["sheet"]]

	spinePrefix = excelVar["spine"]["prefix"]
	spineHostnameCol = excelVar["spine"]["props"]["hostname"]["col"]

	print("spinePrefix = ", inventory_worksheet)

	for row in inventory_worksheet.iter_rows():
		for cell in row:
			print(cell.value)
			if cell.value:
				if eq(cell.coordinate, spineHostnameCol + str(cell.row)):
					p = re.compile(spinePrefix)
					if (p.match(str(cell.value))):
						codi = excelVar["spine"]["props"]["mgmt"]["col"] + str(cell.row)
						mgmtIp = inventory_worksheet[codi].value
						spines_info["hosts"][cell.value] = {"ansible_host": mgmtIp}
	
	return spines_info

def parseLeafInfo(inventory_file, excelVar, leaf_type="L3"):
	'''
	type: options are 'L3' or 'L2'
	'''
	
	workbook = load_workbook(filename=inventory_file, read_only=True, data_only=True)
	leafTypeName = "l3leaf" if leaf_type == "L3" else "l2leaf"
	inventory_worksheet = workbook[excelVar["leaf"]["sheet"]]
	leafs = {"vars": {"type": leafTypeName}, "hosts": {}}
	
	# transform the workbook to a list of dictionaries
	leafPrefix = excelVar["leaf"]["prefix"]
	leafHostnameCol = excelVar["leaf"]["props"]["hostname"]["col"]
	

	for row in inventory_worksheet.iter_rows():
		leaf_info = {}
		for cell in row:
			# print(cell)
			if cell.value:
				if eq(cell.coordinate, leafHostnameCol + str(cell.row)):
					p = re.compile(leafPrefix)
					if (p.match(str(cell.value))):
						hostname = cell.value
						codi = excelVar["leaf"]["props"]["mgmt"]["col"] + str(cell.row)
						mgmtIp = inventory_worksheet[codi].value
      
						leafs["hosts"][cell.value] = {"ansible_host": mgmtIp}


	return leafs

def getServers(fabric_name):
	servers = {"children": {fabric_name + "_L3LEAFS": None, fabric_name + "_L2_LEAFS": None}}
	return servers

def getTenantNetworks(fabric_name):
	tn = {"children": {fabric_name + "_L3LEAFS": None, fabric_name + "_L2LEAFS": None}}
	return tn

def getFabricInventory(inventory_file, fabric_name, excelVar):
	fabric_inventory = {"children":{}}

	# fabric_inventory["children"][fabric_name+"_SUPERSPINES"] = parseSuperSpineInfo(inventory_file)

	fabric_inventory["children"][fabric_name+"_SPINES"] = parseSpineInfo(inventory_file, excelVar)
	
	if parseLeafInfo(inventory_file, excelVar, leaf_type="L3") != None:
		fabric_inventory["children"][fabric_name+"_L3LEAFS"] = parseLeafInfo(inventory_file, excelVar, leaf_type="L3")
		
	fabric_inventory["vars"] = {
		"ansible_connection": "network_cli",
		"ansible_network_os": "eos",
		"ansible_become": True,
		"ansible_user": "ansible",
    "ansible_ssh_pass": "ansible",
		"ansible_become_method": "enable",
		"ansible_httpapi_use_ssl": False,
		"ansible_httpapi_validate_certs": False
	}
	return fabric_inventory

def generateInventory(inventory_file, excelVar):
	fabric_name = getFabricName(inventory_file, excelVar)

	if fabric_name is None:
		return

	inventory = {
		"all": {
			"children": {
				fabric_name: {
					"children": {
						fabric_name + "_FABRIC": {
							"children": { 
								fabric_name + "_SUPERSPINES" : None,
								fabric_name + "_SPINES" : None,
								fabric_name + "_L3LEAFS" : None,
								fabric_name + "_L2LEAFS" : None								
							}
						}
					}					
				}
			}
		}
	}
	# #Add CVP info
	# inventory["all"]["children"]["CVP"] = getCVPInventory(inventory_file)

	#Add Fabric info
	inventory["all"]["children"][fabric_name]["children"][fabric_name + "_FABRIC"] = getFabricInventory(inventory_file, fabric_name, excelVar)

	# inventory = {"all":{"children":{
	#     "CVP": None,
	#     fabric_name: None
	# }}}
	#Add CVP info
	# inventory["all"]["children"]["CVP"] = getCVPInventory(inventory_file)

	# #Add Fabric info
	# inventory["all"]["children"][fabric_name] = getFabricInventory(inventory_file)

	#Add Servers
	inventory["all"]["children"][fabric_name+"_SERVERS"] = getServers(fabric_name)

	#Add Tenant Networks
	inventory["all"]["children"][fabric_name+"_TENANTS_NETWORKS"] = getTenantNetworks(fabric_name)

	return inventory

if __name__ == "__main__":
    generateInventory("PotentialAnsibleCSVTemplate.xlsx")